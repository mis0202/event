import os
import re
from datetime import datetime
from openpyxl import Workbook

# 创建一个新的工作簿
workbook = Workbook()
sheet = workbook.active

# 设置表头
sheet['A1'] = '日期'
sheet['B1'] = '代码行数'

# 遍历代码文件并统计行数
root_dir = r'D:\Projects\djangoProject\event\app01\models.py'

row = 2  # 从第二行开始写入数据

for root, dirs, files in os.walk(root_dir):
    for file in files:
        file_path = os.path.join(root, file)
        date = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d')
        lines_count = 0

        with open(file_path, 'r') as f:
            for line in f:
                # line = line.strip()

                # 检查是否为注释行
                if not re.match(r'^\s*(#|//|/\*|\*)', line):
                    # 检查是否为空行
                    if line != '':
                        lines_count += 1

        # 写入数据到工作表
        sheet.cell(row=row, column=1, value=date)
        sheet.cell(row=row, column=2, value=lines_count)
        row += 1
        print(date)
        print(lines_count)
# 保存工作簿为 Excel 文件
workbook.save('code_stats.xlsx')
